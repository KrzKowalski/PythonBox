import pandas as pd
from openpyxl import load_workbook


def open_rule_file(file_path):
    """Function opens excel workbook and converts it into a DataFrame"""
    rule_wb = load_workbook(file_path, read_only=True)
    ws = rule_wb[rule_wb.sheetnames[0]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row[0:3])
    header = ['RuleGroup', 'RulePriority', 'RuleConstraint']
    rule_df = pd.DataFrame(data, columns=header)
    rule_wb.close()
    print(f"Opening '{file_path}'...")
    return rule_df


def get_rules(rule_dataframe):
    """Function gathers the rules from given DataFrame of mapping file. Converts from strings to a dictionaries."""
    rule_set = []
    for cell in rule_dataframe['RuleConstraint']:
        rule_list = convert_to_list(cell)
        rule_dict = convert_to_dictionary(rule_list)
        rule_set.append(rule_dict)
    print("Preparing rule set...")
    return rule_set


def convert_to_list(rule_string):
    """Function converts the logic string to a list with a keys (in a { }) and a values (in a " ").
    Function removes the elements with comparison operators like: >, <, >=, <="""
    rule_string = rule_string.replace(" ", "").replace("\"", "").replace("\'", "")
    to_replace = ["==", "!=", "in", "not", "and", "or"]
    to_remove = ["<", ">", "<=", ">="]
    for i in to_replace:
        rule_string = rule_string.replace(i, "#")
    rule_string = rule_string.replace("##", "#").split("#")
    for j in rule_string:
        for k in to_remove:
            if k in j:
                rule_string.remove(j)
    for n in rule_string:
        if "(" in n and ")" not in n:
            rule_string[rule_string.index(n)] = n.replace("(", "")
        if ")" in n and "(" not in n:
            rule_string[rule_string.index(n)] = n.replace(")", "")
    # print(rule_string)
    return rule_string


def convert_to_dictionary(rule_list):
    """Function converts a given rule expressed as a list to a dictionary."""
    key_index = [i for i in range(len(rule_list)) if "{" in rule_list[i]]
    rule_dict = {}
    temp_list = []
    for elem in rule_list:
        elem = elem.replace("{", "").replace("}", "")
        temp_list.append(elem)
    rule_list = temp_list
    for position in key_index:
        if rule_list[position] not in rule_dict:
            rule_dict[rule_list[position]] = [rule_list[position + 1]]
        else:
            rule_dict[rule_list[position]].append(rule_list[position + 1])
    return rule_dict


def perform_checks(input_df, rule_collection):
    """Function performs checks - compares the input file (converted to a DataFrame) with a rule collection. It returns
    a list of tuples that contain a index and a potential match information"""
    print("Comparing the rule set with a input file...")
    headers = [_ for _ in input_df]
    rule_count = len(input_df.index)
    match_list = []
    for rule_as_list in rule_collection:
        for rule_key in rule_as_list:
            header_to_check = headers.index(rule_key) + 1
            for rule_value in rule_as_list[rule_key]:
                for input_line in input_df.itertuples(name='row'):
                    if input_line[header_to_check].replace(" ", "") in rule_value:
                        report_tuple = (rule_collection.index(rule_as_list), f"Match on {rule_as_list[rule_key][0]}")
                        match_list.append(report_tuple)

    print(f"Matches found: {len(match_list)}. Total rules checked: {len(rule_collection)}. Input positions checked: {rule_count}.")
    return match_list


def generate_report(report_file_name, rule_set_df, match_records):
    """Function generates a report that contains information
    if the rule in a mapping file has a match in a input file."""
    report_df = rule_set_df[rule_set_df.keys()[0:3]]
    report_df.insert(3, 'ReportComment', 'No match found')
    temp_dict = report_df.to_dict(orient='index')
    for line in temp_dict:
        for match in match_records:
            if line == match[0]:
                temp_dict[line]['ReportComment'] = match[1]
    final_report_df = pd.DataFrame.from_dict(temp_dict, orient='index')
    final_report_df.to_excel(report_file_name, index=False)
    print(f"Report generated: {report_file_name}")


#
# rule_df = open_rule_file('rule_file.xlsx')
# rule_set = get_rules(rule_df)
# input_file = 'input_file.csv'
# input_df = pd.read_csv(input_file, sep=";")
# print(f"Opening {input_file}")
# match_list = perform_checks(input_df, rule_set)
# generate_report('report_test1_output.xlsx', rule_df, match_list)
#

# BELOW IS FOR SS:
print("==========================================================")
rule_df = open_rule_file('Enrich.xlsx')
rule_set = get_rules(rule_df)
input_file = 'testing_input.csv'
input_df = pd.read_csv(input_file, low_memory=False)
print(f"Opening {input_file}")
match_list = perform_checks(input_df, rule_set)
generate_report('report_test1_output.xlsx', rule_df, match_list)
print("==========================================================")